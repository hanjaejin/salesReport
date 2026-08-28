"""여러 매장 보고서(xlsx) 생성 — 부록 B.11.

중간 관리자가 팀장에게 **건네줄 파일**이다. 매장별 일일 보고서(명세 9장)의
그룹판이며, 같은 원칙을 그대로 따른다.

- 화면과 **같은 원천**에서 만든다 — 저장된 그룹 요약과 마트다. 숫자가 어긋날 수 없다.
- 새로 계산하지 않는다. 합계·비중은 배치가, 기간 집계는 DB가 만든 값을 옮겨 적을 뿐이다.
- 전문용어를 쓰지 않고 목업 배지를 넣는다 (명세 14장).
- 메모리에서 만들어 돌려준다 — 클라우드에 쓰기 권한이 없어도 내려받을 수 있다.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import Engine

from src.common.dateutil import parse_date
from src.common.logger import get_logger

logger = get_logger(__name__)

#: 시트 구성 (부록 B.11)
SHEET_SUMMARY = "요약"
SHEET_STORES = "매장별"
SHEET_PERIOD = "기간"
SHEET_HOURLY = "시간대"

#: 머리글 스타일 — 매장 보고서와 같은 남색을 쓴다.
_HEADER_FILL = PatternFill("solid", fgColor="22324A")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True)
_MONEY_FORMAT = "#,##0"

#: 상태 코드를 사람 말로 (부록 B.6과 같은 어휘를 쓴다).
_STATUS_LABELS: dict[str, str] = {
    "STOCK": "재고 주의",
    "PEAK": "시간대 쏠림",
    "CALM": "조용한 날",
}

MOCKUP_NOTE = "목업 데이터 — 실제 매출이 아닙니다."
NO_ATTENTION = "특별히 먼저 볼 매장 없음"
NO_BASELINE = "-"


def _style_header(sheet: Worksheet, row: int, width: int) -> None:
    """머리글 행에 스타일을 입힌다.

    Args:
        sheet: 대상 시트.
        row: 머리글 행 번호.
        width: 열 개수.
    """
    for column in range(1, width + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _fit_columns(sheet: Worksheet, widths: tuple[int, ...]) -> None:
    """열 너비를 지정한다 — 열어 보자마자 읽히도록.

    Args:
        sheet: 대상 시트.
        widths: 열 너비 목록.
    """
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_summary(sheet: Worksheet, payload: dict, period: dict) -> None:
    """요약 시트를 쓴다.

    Args:
        sheet: 대상 시트.
        payload: 저장된 그룹 요약 JSON.
        period: 기간 집계 결과.
    """
    saledate = parse_date(payload["saledate"]).strftime("%Y-%m-%d")

    sheet["A1"] = "여러 매장 보고"
    sheet["A1"].font = _TITLE_FONT
    sheet["A2"] = f"기준일 {saledate} · {payload['store_count']}개 매장"
    sheet["A3"] = MOCKUP_NOTE

    sheet.append([])
    sheet.append(["항목", "값"])
    _style_header(sheet, sheet.max_row, 2)

    rows: list[tuple[str, object]] = [
        ("합계 매출", payload["total_sale_amt"]),
        ("합계 손님", payload["total_deal_cnt"]),
        ("1인당", payload["group_avg_ticket"]),
        (f"최근 {period['days']}일 매출", period["sale_amt"]),
        (f"최근 {period['days']}일 손님", period["deal_cnt"]),
        (
            f"지난 {period['days']}일 대비",
            NO_BASELINE if period["prev_diff_pct"] is None else f"{period['prev_diff_pct']}%",
        ),
        ("먼저 볼 매장", payload["attention_line"]),
    ]
    for label, value in rows:
        sheet.append([label, value])
        if isinstance(value, int):
            sheet.cell(row=sheet.max_row, column=2).number_format = _MONEY_FORMAT

    _fit_columns(sheet, (22, 34))


def _write_stores(sheet: Worksheet, payload: dict) -> None:
    """매장별 시트를 쓴다.

    Args:
        sheet: 대상 시트.
        payload: 저장된 그룹 요약 JSON.
    """
    sheet.append(["매장", "매출", "손님", "1인당", "비중", "평소 같은 요일 대비", "상태"])
    _style_header(sheet, 1, 7)

    for row in payload["stores"]:
        sheet.append(
            [
                row["dept_nm"],
                row["sale_amt"],
                row["deal_cnt"],
                row["avg_ticket"],
                f"{row['share_pct']}%",
                NO_BASELINE
                if not row["dow_baseline_available"]
                else f"{row['dow_diff_pct']}%",
                f"{_STATUS_LABELS[row['status']]} — {row['status_text']}",
            ]
        )
        for column in (2, 3, 4):
            sheet.cell(row=sheet.max_row, column=column).number_format = _MONEY_FORMAT

    _fit_columns(sheet, (18, 14, 10, 10, 8, 20, 26))


def _write_period(sheet: Worksheet, period: dict) -> None:
    """기간 시트를 쓴다.

    Args:
        sheet: 대상 시트.
        period: 기간 집계 결과.
    """
    from_date = parse_date(period["from_date"]).strftime("%Y-%m-%d")
    to_date = parse_date(period["to_date"]).strftime("%Y-%m-%d")

    sheet["A1"] = f"최근 {period['days']}일 매장별 합계"
    sheet["A1"].font = _TITLE_FONT
    sheet["A2"] = f"집계 기간 {from_date} ~ {to_date}"

    sheet.append([])
    sheet.append(["매장", "매출", "손님"])
    _style_header(sheet, sheet.max_row, 3)

    for row in period["stores"]:
        sheet.append([row["dept_nm"], row["sale_amt"], row["deal_cnt"]])
        for column in (2, 3):
            sheet.cell(row=sheet.max_row, column=column).number_format = _MONEY_FORMAT

    _fit_columns(sheet, (18, 16, 12))


def _write_hourly(sheet: Worksheet, hourly: object) -> None:
    """시간대 시트를 쓴다 (매장 × 시간대).

    Args:
        sheet: 대상 시트.
        hourly: 시간을 인덱스로, 매장명을 열로 갖는 프레임.
    """
    sheet.append(["시간", *hourly.columns])  # type: ignore[attr-defined]
    _style_header(sheet, 1, len(hourly.columns) + 1)  # type: ignore[attr-defined]

    for label, values in hourly.iterrows():  # type: ignore[attr-defined]
        sheet.append([label, *(int(value) for value in values)])
        for column in range(2, len(values) + 2):
            sheet.cell(row=sheet.max_row, column=column).number_format = _MONEY_FORMAT

    _fit_columns(sheet, (10, *(16,) * len(hourly.columns)))  # type: ignore[attr-defined]


def build_workbook(engine: Engine, saledate: str) -> Workbook:
    """여러 매장 보고서를 만든다 (부록 B.11).

    Args:
        engine: 대상 엔진.
        saledate: ``YYYYMMDD``.

    Returns:
        openpyxl 워크북.

    Raises:
        LookupError: 해당 일자의 그룹 요약이 없을 때.
    """
    # 화면과 같은 로더를 쓴다 — 원천이 같아야 숫자가 어긋나지 않는다 (흐름도 FLOW 06).
    from src.app.main import load_group_briefing, load_group_hourly, load_period_summary

    payload = load_group_briefing(engine, saledate)
    if payload is None:
        raise LookupError(f"해당 일자의 여러 매장 요약이 없습니다: {saledate}")

    period = load_period_summary(engine, saledate)
    hourly = load_group_hourly(engine, saledate)

    workbook = Workbook()
    summary = workbook.active
    summary.title = SHEET_SUMMARY

    _write_summary(summary, payload, period)
    _write_stores(workbook.create_sheet(SHEET_STORES), payload)
    _write_period(workbook.create_sheet(SHEET_PERIOD), period)
    _write_hourly(workbook.create_sheet(SHEET_HOURLY), hourly)

    return workbook


def report_bytes(engine: Engine, saledate: str) -> bytes:
    """보고서를 메모리에서 바로 바이트로 만든다 (화면 내려받기용).

    Args:
        engine: 대상 엔진.
        saledate: ``YYYYMMDD``.

    Returns:
        xlsx 바이트열.

    Raises:
        LookupError: 해당 일자의 그룹 요약이 없을 때.
    """
    buffer = io.BytesIO()
    build_workbook(engine, saledate).save(buffer)
    logger.info("여러 매장 보고서 생성 완료: %s", saledate)
    return buffer.getvalue()
