"""일일 보고서(xlsx) 생성 — 명세 9장.

화면과 **같은 마트**에서 만든다. 원천이 같으므로 화면 숫자와 보고서 숫자가
어긋날 수 없다 (흐름도 FLOW 06).

표기 반올림은 브리핑과 같은 규칙을 쓴다 — 금액은 정수 원, 1인당 구매액은 정수 원,
%는 소수 1자리 (명세 7.4).
"""

from __future__ import annotations

import io
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import Engine, select

from src.common.dateutil import parse_date
from src.common.logger import get_logger
from src.load import schema

logger = get_logger(__name__)

#: 보고서 시트 구성 (명세 9장: "요약+TOP5+시간대 시트")
SHEET_SUMMARY = "요약"
SHEET_TOP5 = "TOP5"
SHEET_HOURLY = "시간대"
SHEET_STOCK = "재고"

#: 머리글 스타일
_HEADER_FILL = PatternFill("solid", fgColor="22324A")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True)
_MONEY_FORMAT = "#,##0"


def report_filename(dept_nm: str, saledate: str) -> str:
    """보고서 파일명을 만든다 (명세 9장: ``일일보고_{점포명}_{일자}.xlsx``).

    Args:
        dept_nm: 점포명.
        saledate: ``YYYYMMDD``.

    Returns:
        파일명.
    """
    return f"일일보고_{dept_nm}_{saledate}.xlsx"


def fetch_report_data(
    engine: Engine, saledate: str, dept_cd: str
) -> tuple[dict, pd.DataFrame, pd.DataFrame]:
    """보고서에 쓸 자료를 마트와 브리핑에서 읽는다.

    xlsx 보고서와 정적 HTML 스냅샷이 **같은 값**을 쓰도록 공개한다 —
    두 산출물의 숫자가 갈라질 여지를 없앤다.

    Args:
        engine: 대상 엔진.
        saledate: ``YYYYMMDD``.
        dept_cd: 점포코드.

    Returns:
        ``(계산 JSON, 상품 상위 프레임, 시간대 프레임)``.

    Raises:
        LookupError: 해당 일자·점포의 브리핑이 없을 때.
    """
    item = schema.MART_DAY_STORE_ITEM
    hour = schema.MART_HOUR_STORE

    with engine.connect() as connection:
        raw = connection.execute(
            select(schema.BRIEFING_DAILY.c.PAYLOAD_JSON).where(
                schema.BRIEFING_DAILY.c.SALEDATE == saledate,
                schema.BRIEFING_DAILY.c.DEPT_CD == dept_cd,
            )
        ).scalar_one_or_none()

        if raw is None:
            raise LookupError(f"브리핑이 없습니다: {saledate} / {dept_cd}")

        top_items = pd.read_sql(
            select(
                item.c.GOODS_NM, item.c.ITEM_HEAD_NM, item.c.SALE_AMT, item.c.QTY
            )
            .where(item.c.SALEDATE == saledate, item.c.DEPT_CD == dept_cd)
            .order_by(item.c.SALE_AMT.desc())
            .limit(5),
            connection,
        )
        hourly = pd.read_sql(
            select(hour.c.HOUR, hour.c.SALE_AMT, hour.c.DEAL_CNT)
            .where(hour.c.SALEDATE == saledate, hour.c.DEPT_CD == dept_cd)
            .order_by(hour.c.HOUR),
            connection,
        )

    return json.loads(raw), top_items, hourly


def _style_header(sheet: Worksheet, row: int, width: int) -> None:
    """머리글 행에 배경·글꼴을 입힌다.

    Args:
        sheet: 대상 시트.
        row: 머리글 행 번호.
        width: 머리글 칸 수.
    """
    for column in range(1, width + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(sheet: Worksheet, widths: dict[int, int]) -> None:
    """열 너비를 지정한다.

    Args:
        sheet: 대상 시트.
        widths: 열 번호 → 너비.
    """
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_summary(sheet: Worksheet, payload: dict) -> None:
    """요약 시트를 쓴다 — 매출·손님 수·1인당 구매액과 그날의 브리핑 3줄.

    Args:
        sheet: 대상 시트.
        payload: 계산 JSON.
    """
    saledate = parse_date(payload["saledate"]).strftime("%Y-%m-%d")

    sheet["A1"] = f"{payload['dept_nm']} 일일 보고"
    sheet["A1"].font = _TITLE_FONT
    sheet["A2"] = f"기준일 {saledate} ({payload['dow_name']})"
    sheet["A3"] = "목업 데이터 — 실제 매출이 아닙니다."

    sheet.append([])
    sheet.append(["항목", "값", "그저께 대비"])
    _style_header(sheet, sheet.max_row, 3)

    rows = [
        ("매출", payload["sale_amt"], payload["prev_diff_pct"]),
        ("손님 수", payload["deal_cnt"], payload["cnt_diff_pct"]),
        ("1인당 구매액", payload["avg_ticket"], payload["ticket_diff_pct"]),
    ]
    for label, value, diff in rows:
        sheet.append([label, value, "-" if diff is None else f"{diff}%"])
        sheet.cell(row=sheet.max_row, column=2).number_format = _MONEY_FORMAT

    sheet.append([])
    sheet.append([f"평소 {payload['dow_name']} 평균", payload["dow_baseline_amt"], ""])
    sheet.cell(row=sheet.max_row, column=2).number_format = _MONEY_FORMAT

    sheet.append([])
    sheet.append(["오늘의 브리핑"])
    sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
    for line in payload["briefing_lines"]:
        sheet.append([line])

    _autosize(sheet, {1: 46, 2: 16, 3: 14})


def _write_top5(sheet: Worksheet, top_items: pd.DataFrame) -> None:
    """TOP5 시트를 쓴다.

    Args:
        sheet: 대상 시트.
        top_items: 매출 상위 상품 프레임.
    """
    sheet.append(["순위", "상품명", "분류", "매출", "판매수량"])
    _style_header(sheet, 1, 5)

    for rank, row in enumerate(top_items.itertuples(index=False), start=1):
        sheet.append([rank, row.GOODS_NM, row.ITEM_HEAD_NM, int(row.SALE_AMT), int(row.QTY)])
        sheet.cell(row=sheet.max_row, column=4).number_format = _MONEY_FORMAT

    _autosize(sheet, {1: 6, 2: 34, 3: 14, 4: 14, 5: 10})


def _write_hourly(sheet: Worksheet, hourly: pd.DataFrame) -> None:
    """시간대 시트를 쓴다.

    Args:
        sheet: 대상 시트.
        hourly: 시간대 마트 프레임.
    """
    sheet.append(["시간대", "매출", "손님 수"])
    _style_header(sheet, 1, 3)

    for row in hourly.itertuples(index=False):
        sheet.append([f"{row.HOUR}시", int(row.SALE_AMT), int(row.DEAL_CNT)])
        sheet.cell(row=sheet.max_row, column=2).number_format = _MONEY_FORMAT

    _autosize(sheet, {1: 12, 2: 16, 3: 10})


def _write_stock(sheet: Worksheet, payload: dict) -> None:
    """재고 시트를 쓴다 — 곧 떨어질 수 있는 상품 (부록 A.7).

    **권고발주수량은 넣지 않는다.** 본 명세 7.4가 발주 수량 제시를 금지한다.

    Args:
        sheet: 대상 시트.
        payload: 계산 JSON.
    """
    sheet.append(["상품", "남은 재고", "하루 평균 판매"])
    _style_header(sheet, 1, 3)

    items = payload.get("stock_risk", {}).get("items", [])
    if not items:
        sheet.append(["지금은 부족한 상품이 없어요", "", ""])
    for item in items:
        sheet.append([item["goods_nm"], item["stock_qty"], item["sale_average_qty"]])

    _autosize(sheet, {1: 34, 2: 12, 3: 14})


def build_workbook(engine: Engine, saledate: str, dept_cd: str) -> Workbook:
    """일일 보고서 워크북을 만든다 (명세 9장).

    Args:
        engine: 대상 엔진.
        saledate: ``YYYYMMDD``.
        dept_cd: 점포코드.

    Returns:
        요약·TOP5·시간대 3개 시트를 가진 워크북.

    Raises:
        LookupError: 해당 일자·점포의 브리핑이 없을 때.
    """
    payload, top_items, hourly = fetch_report_data(engine, saledate, dept_cd)

    workbook = Workbook()
    summary = workbook.active
    summary.title = SHEET_SUMMARY

    _write_summary(summary, payload)
    _write_top5(workbook.create_sheet(SHEET_TOP5), top_items)
    _write_hourly(workbook.create_sheet(SHEET_HOURLY), hourly)
    _write_stock(workbook.create_sheet(SHEET_STOCK), payload)

    return workbook


def report_bytes(engine: Engine, saledate: str, dept_cd: str) -> bytes:
    """보고서를 메모리에서 바로 바이트로 만든다 (화면 내려받기용).

    임시 파일을 만들지 않으므로 클라우드 배포에서도 쓰기 권한이 필요 없다.

    Args:
        engine: 대상 엔진.
        saledate: ``YYYYMMDD``.
        dept_cd: 점포코드.

    Returns:
        xlsx 바이트열.
    """
    buffer = io.BytesIO()
    build_workbook(engine, saledate, dept_cd).save(buffer)
    return buffer.getvalue()


def write_daily_report(
    engine: Engine, saledate: str, dept_cd: str, out_dir: Path
) -> Path:
    """보고서를 파일로 저장한다 (배치·검증용).

    Args:
        engine: 대상 엔진.
        saledate: ``YYYYMMDD``.
        dept_cd: 점포코드.
        out_dir: 저장할 디렉토리. 없으면 만든다.

    Returns:
        저장한 파일 경로.

    Raises:
        LookupError: 해당 일자·점포의 브리핑이 없을 때.
    """
    payload, _, _ = fetch_report_data(engine, saledate, dept_cd)

    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / report_filename(payload["dept_nm"], saledate)
    build_workbook(engine, saledate, dept_cd).save(path)

    logger.info("보고서 생성: %s", path)
    return path
