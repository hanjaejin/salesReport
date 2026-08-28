"""일일 보고서 검증 — 명세 9장·10장 ``test_report_xlsx_created``.

보고서는 제출물이다. 화면 숫자와 어긋나면 그 자리에서 신뢰를 잃으므로
"파일이 만들어졌다"에 그치지 않고 **내용이 마트와 같은지**까지 본다.
"""

from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import Engine, select

from src.common.config import get_engine
from src.extract.sample import SampleExtractor
from src.load import pipeline, schema
from src.report import daily_report

SALEDATE = "20260703"
DEPT_CD = "901002"
DEPT_NM = "동부역 중형점"


@pytest.fixture(scope="module")
def built_engine(tmp_path_factory: pytest.TempPathFactory) -> Engine:
    """보고서를 만들 수 있는 상태의 엔진 (모듈 1회)."""
    engine = get_engine(tmp_path_factory.mktemp("report") / "report.db")
    pipeline.load_period(SampleExtractor(), "20260701", "20260705", engine=engine)
    return engine


def _payload(engine: Engine, saledate: str, dept_cd: str) -> dict:
    """저장된 브리핑 JSON을 읽는다.

    Args:
        engine: 대상 엔진.
        saledate: ``YYYYMMDD``.
        dept_cd: 점포코드.

    Returns:
        계산 JSON.
    """
    with engine.connect() as connection:
        raw = connection.execute(
            select(schema.BRIEFING_DAILY.c.PAYLOAD_JSON).where(
                schema.BRIEFING_DAILY.c.SALEDATE == saledate,
                schema.BRIEFING_DAILY.c.DEPT_CD == dept_cd,
            )
        ).scalar_one()
    return json.loads(raw)


def test_report_xlsx_created(built_engine: Engine, tmp_path: Path) -> None:
    """명세 10장: 보고서 파일이 생성되고 세 시트가 존재한다."""
    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)

    assert path.exists()
    assert path.stat().st_size > 0

    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        daily_report.SHEET_SUMMARY,
        daily_report.SHEET_TOP5,
        daily_report.SHEET_HOURLY,
        daily_report.SHEET_STOCK,
    ]


def test_report_filename_matches_spec() -> None:
    """명세 9장: 파일명은 ``일일보고_{점포명}_{일자}.xlsx``."""
    assert (
        daily_report.report_filename(DEPT_NM, SALEDATE)
        == f"일일보고_{DEPT_NM}_{SALEDATE}.xlsx"
    )


def test_report_file_is_named_after_store(built_engine: Engine, tmp_path: Path) -> None:
    """저장된 파일명이 점포명·일자를 담는다."""
    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)

    assert path.name == f"일일보고_{DEPT_NM}_{SALEDATE}.xlsx"


def test_summary_numbers_match_briefing(built_engine: Engine, tmp_path: Path) -> None:
    """요약 시트의 수치가 브리핑 JSON과 같다 (화면·보고서 불일치 원천 차단)."""
    payload = _payload(built_engine, SALEDATE, DEPT_CD)
    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)

    sheet = load_workbook(path)[daily_report.SHEET_SUMMARY]
    values = {row[0]: row[1] for row in sheet.iter_rows(values_only=True) if row[0]}

    assert values["매출"] == payload["sale_amt"]
    assert values["손님 수"] == payload["deal_cnt"]
    assert values["1인당 구매액"] == payload["avg_ticket"]


def test_summary_carries_briefing_lines(built_engine: Engine, tmp_path: Path) -> None:
    """요약 시트에 그날의 브리핑 3줄이 그대로 담긴다 (재작성 없음)."""
    payload = _payload(built_engine, SALEDATE, DEPT_CD)
    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)

    sheet = load_workbook(path)[daily_report.SHEET_SUMMARY]
    texts = {row[0] for row in sheet.iter_rows(values_only=True) if isinstance(row[0], str)}

    for line in payload["briefing_lines"]:
        assert line in texts


def test_top5_matches_mart(built_engine: Engine, tmp_path: Path) -> None:
    """TOP5 시트가 상품 마트의 상위 5개와 일치한다."""
    item = schema.MART_DAY_STORE_ITEM
    with built_engine.connect() as connection:
        expected = connection.execute(
            select(item.c.GOODS_NM, item.c.SALE_AMT, item.c.QTY)
            .where(item.c.SALEDATE == SALEDATE, item.c.DEPT_CD == DEPT_CD)
            .order_by(item.c.SALE_AMT.desc())
            .limit(5)
        ).all()

    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)
    sheet = load_workbook(path)[daily_report.SHEET_TOP5]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    assert len(rows) == len(expected)
    for row, (goods_nm, sale_amt, qty) in zip(rows, expected, strict=True):
        assert row[1] == goods_nm
        assert row[3] == sale_amt
        assert row[4] == qty


def test_hourly_matches_mart(built_engine: Engine, tmp_path: Path) -> None:
    """시간대 시트가 시간대 마트와 일치한다."""
    hour = schema.MART_HOUR_STORE
    with built_engine.connect() as connection:
        expected = connection.execute(
            select(hour.c.HOUR, hour.c.SALE_AMT)
            .where(hour.c.SALEDATE == SALEDATE, hour.c.DEPT_CD == DEPT_CD)
            .order_by(hour.c.HOUR)
        ).all()

    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)
    sheet = load_workbook(path)[daily_report.SHEET_HOURLY]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    assert len(rows) == len(expected)
    for row, (hour_text, sale_amt) in zip(rows, expected, strict=True):
        assert row[0] == f"{hour_text}시"
        assert row[1] == sale_amt


def test_report_bytes_is_valid_workbook(built_engine: Engine) -> None:
    """메모리 생성본도 같은 시트 구성을 갖는다 (화면 내려받기 경로)."""
    import io

    payload = daily_report.report_bytes(built_engine, SALEDATE, DEPT_CD)

    assert payload[:2] == b"PK", "xlsx(zip) 시그니처가 아니다"
    workbook = load_workbook(io.BytesIO(payload))
    assert len(workbook.sheetnames) == 4


def test_missing_day_raises_lookup_error(built_engine: Engine, tmp_path: Path) -> None:
    """브리핑이 없는 날짜는 조용히 빈 파일을 만들지 않고 실패한다."""
    with pytest.raises(LookupError, match="브리핑이 없습니다"):
        daily_report.write_daily_report(built_engine, "20991231", DEPT_CD, tmp_path)


def test_report_has_mockup_notice(built_engine: Engine, tmp_path: Path) -> None:
    """보고서에도 목업 고지가 있다 (화면 배지와 같은 정직성)."""
    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)

    sheet = load_workbook(path)[daily_report.SHEET_SUMMARY]
    texts = " ".join(
        str(row[0]) for row in sheet.iter_rows(values_only=True) if isinstance(row[0], str)
    )

    assert "목업" in texts


def test_report_uses_no_jargon(built_engine: Engine, tmp_path: Path) -> None:
    """명세 14장: 보고서에도 전문용어를 쓰지 않는다."""
    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)

    workbook = load_workbook(path)
    texts = " ".join(
        str(cell)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if isinstance(cell, str)
    )

    for word in ("객단가", "증감률", "AI", "LLM"):
        assert word not in texts, f"금지 용어 '{word}' 가 보고서에 있다"


# --- 부록 A.7: 재고 시트 -----------------------------------------------------


def test_stock_sheet_matches_briefing(built_engine: Engine, tmp_path: Path) -> None:
    """재고 시트가 브리핑의 위험 품목과 일치한다."""
    payload = _payload(built_engine, SALEDATE, DEPT_CD)
    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)

    sheet = load_workbook(path)[daily_report.SHEET_STOCK]
    rows = [r for r in sheet.iter_rows(min_row=2, values_only=True) if r[0]]
    items = payload["stock_risk"]["items"]

    if items:
        assert [r[0] for r in rows] == [item["goods_nm"] for item in items]
        assert [r[1] for r in rows] == [item["stock_qty"] for item in items]
    else:
        assert rows[0][0] == "지금은 부족한 상품이 없어요"


def test_stock_sheet_omits_order_quantity(built_engine: Engine, tmp_path: Path) -> None:
    """명세 7.4: 보고서에도 발주 수량을 제시하지 않는다."""
    path = daily_report.write_daily_report(built_engine, SALEDATE, DEPT_CD, tmp_path)

    sheet = load_workbook(path)[daily_report.SHEET_STOCK]
    headers = next(sheet.iter_rows(values_only=True))

    assert headers == ("상품", "남은 재고", "하루 평균 판매")
    texts = " ".join(
        str(c) for row in sheet.iter_rows(values_only=True) for c in row if isinstance(c, str)
    )
    for forbidden in ("발주", "권고", "적정재고", "리드타임"):
        assert forbidden not in texts


# --- 부록 B.11: 여러 매장 보고서 ---------------------------------------------


def test_group_report_has_expected_sheets(built_engine: Engine) -> None:
    """부록 B.11: 시트 4종."""
    from openpyxl import load_workbook

    from src.report import group_report

    book = load_workbook(io.BytesIO(group_report.report_bytes(built_engine, SALEDATE)))

    assert book.sheetnames == ["요약", "매장별", "기간", "시간대"]


def test_group_report_numbers_match_screen(built_engine: Engine) -> None:
    """부록 B.11: 보고서 숫자가 화면 숫자와 같다 — 같은 원천에서 만든다."""
    from openpyxl import load_workbook

    from src.app import main
    from src.report import group_report

    payload = main.load_group_briefing(built_engine, SALEDATE)
    book = load_workbook(io.BytesIO(group_report.report_bytes(built_engine, SALEDATE)))
    summary = {
        row[0]: row[1]
        for row in book[group_report.SHEET_SUMMARY].iter_rows(values_only=True)
        if row[0]
    }

    assert summary["합계 매출"] == payload["total_sale_amt"]
    assert summary["합계 손님"] == payload["total_deal_cnt"]
    assert summary["1인당"] == payload["group_avg_ticket"]


def test_group_report_lists_every_store(built_engine: Engine) -> None:
    """매장별 시트에 세 매장이 모두 들어간다."""
    from openpyxl import load_workbook

    from src.report import group_report

    book = load_workbook(io.BytesIO(group_report.report_bytes(built_engine, SALEDATE)))
    text = " ".join(
        str(cell)
        for row in book["매장별"].iter_rows(values_only=True)
        for cell in row
        if isinstance(cell, str)
    )

    for name in ("중앙역 대형점", "동부역 중형점", "간이역 소형점"):
        assert name in text


def test_group_report_shows_no_jargon(built_engine: Engine) -> None:
    """부록 B.12: 보고서에도 전문용어가 없다 (명세 14장)."""
    from openpyxl import load_workbook

    from src.report import group_report

    book = load_workbook(io.BytesIO(group_report.report_bytes(built_engine, SALEDATE)))
    text = " ".join(
        str(cell)
        for sheet in book.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if isinstance(cell, str)
    )

    for word in ("객단가", "증감률", "AI", "LLM", "분석", "예측", "결품", "소진일수"):
        assert word not in text
    assert "목업" in text


def test_group_report_missing_day_raises(built_engine: Engine) -> None:
    """없는 날짜는 조용히 빈 파일을 만들지 않고 알린다 (빈 except 금지)."""
    import pytest as _pytest

    from src.report import group_report

    with _pytest.raises(LookupError):
        group_report.report_bytes(built_engine, "20991231")
